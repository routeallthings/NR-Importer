#!/usr/bin/env python
'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
GIT (If missing any of the modules)
netaddr
requests

---VERSION---
VERSION 1.0
Currently Implemented Features
-


Features planned in the near future


'''

'''IMPORT MODULES'''
import getpass
import os
import sys
#
try:
	import netaddr
	from netaddr import *
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('netaddr module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in requestsinstallstatus or "y" in requestsinstallstatus or "yes" in requestsinstallstatus or "Yes" in requestsinstallstatus or "YES" in requestsinstallstatus:
		os.system('python -m pip install netaddr')
		import netaddr
		from netaddr import *
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of netaddr. Please install manually and retry"
		sys.exit()
import xml.etree.cElementTree as ET
import xml.dom.minidom
#
try:
	import requests
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('Requests module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in requestsinstallstatus or "y" in requestsinstallstatus or "yes" in requestsinstallstatus or "Yes" in requestsinstallstatus or "YES" in requestsinstallstatus:
		os.system('python -m pip install requests')
		import requests
	else:
		print "You selected an option other than yes. Please be aware that this script requires the use of Requests. Please install manually and retry"
		sys.exit()
from requests.auth import HTTPBasicAuth
#
try:
	from openpyxl import load_workbook
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
#	
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper
#
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()
# Export to file or direct import
ImportTypeq = raw_input('Do you want to (e)xport to a file or (i)mport directly to the RESTAPI (e.g. E or I)?:')
if 'e' in ImportTypeq.lower():
	ExportLocation = raw_input('Folder path to export the XML files (e.g. C:\\Python27):')
	if ExportLocation == '':
		ExportLocation = 'C:/Python27'
	if not ExportLocation.endswith('/'):
		ExportLocation = ExportLocation + '/'
	if not os.path.exists(ExportLocation):
		os.makedirs(ExportLocation)
policyq = raw_input('Do you want to import the Template Policy (Y/N)?:')
if policyq == '':
	policyq = 'Y'
scopeq = raw_input('Do you want to import the DHCP Scopes (Y/N)?:')
if scopeq == '':
	scopeq = 'Y'
		
# Connect to Network Registrar
if 'i' in ImportTypeq.lower():
	NRURLq = raw_input('What is the IP Address of the NR server (e.g. 10.1.1.1)? ')
	NRPORTq = raw_input('What is the port for connecting to the REST service (e.g. 8080)?:')
	NRUSERq = raw_input('What is the username of the login for the NR server?:')
	NRPASSWORDq = getpass.getpass('What is the password for the NR user?:')
	if NRPORTq == '':
				NRPORTq = '8080'
	nrurl = "http://" + NRURLq + ':' + NRPORTq + "/web-services/rest"
	# NR-BaseURL
	nrpolicyurl = nrurl + '/resource/Policy'
	nrscopeurl = nrurl + '/resource/Scope'
	headers = {'Content-Type': 'application/xml'}
# Import Excel File
excelfilelocation = raw_input('Excel file to load the data from (e.g. C:/Python27/NRDATA.xlsx):')
if excelfilelocation == '':
	excelfilelocation = 'C:/Python27/NRDATA.xlsx'
try:
	excelfiledata = load_workbook(excelfilelocation)
	exceldatasheet = excelfiledata.worksheets[0]
	excelq1 = raw_input('What is the name of the column you want to match on (e.g. ZoneName)?:')
	if excelq1 == '':
		excelq1 = 'ZoneName'
	excelq2 = raw_input('Do you want to do a partial import (y/n)?:')
	if 'n' in excelq2.lower():
		excelq3 = '10000'
		excelq3 = int(excelq3)
	if 'y' in excelq2.lower():
		excelq3 = raw_input('How many rows do you want to import (e.g. 3)?:')
		excelq3 = int(excelq3)
except:
	print 'Error in loading the excel file for the data input. Please enter a good path for the file.'

# Start of Loop
print 'Starting import of data'
countloop = 0
for rowdata in xlhelper.sheet_to_dict(excelfilelocation,'FinalData'):
	countloop = countloop + 1
	if countloop > int(excelq3):
		break
	try:
		ZoneName = rowdata.get(excelq1).encode('utf-8')
		if 'i' in ImportTypeq.lower():
			ZonePolicyURL = nrpolicyurl + '/' + ZoneName
			ZoneScopeURL = nrscopeurl + '/' + ZoneName
		if "y" in policyq.lower():
			## ZONE POLICY START ##
			policyroot = ET.Element("Policy")
			# Policy - GracePeriod
			try:
				gracePeriod = rowdata.get('GracePeriod').encode('utf-8')
			except:
				gracePeriod = rowdata.get('GracePeriod')
				gracePeriod = str(gracePeriod)
			ET.SubElement(policyroot, "gracePeriod").text = gracePeriod
			# Policy - Name
			ET.SubElement(policyroot, "name").text = ZoneName
			# Policy - OfferTimeout
			try:
				offerTimeout = rowdata.get('offerTimeout').encode('utf-8')
			except:
				offerTimeout = rowdata.get('offerTimeout')
				offerTimeout = str(offerTimeout)
			ET.SubElement(policyroot, "offerTimeout").text = offerTimeout
			# Policy - OptionList
			optionlist = ET.SubElement(policyroot, "optionList")
			# Policy - OptionItems - Gateway
			try:
				Gateway = rowdata.get('Gateway').encode('utf-8')
			except:
				Gateway = rowdata.get('Gateway')
				Gateway = str(Gateway)
			gatewayoption = ET.SubElement (optionlist, "OptionItem")
			ET.SubElement(gatewayoption, "number").text = '3'
			ET.SubElement(gatewayoption, "value").text = Gateway
			ET.SubElement(gatewayoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Policy - OptionItems - DNSServer
			try:
				DNS1 = rowdata.get('DNS1').encode('utf-8')
				DNS2 = rowdata.get('DNS2').encode('utf-8')
			except:
				DNS1 = rowdata.get('DNS1')
				DNS1 = str(DNS1)
				DNS2 = rowdata.get('DNS2')
				DNS2 = str(DNS2)
			dnsoption = ET.SubElement (optionlist, "OptionItem")
			ET.SubElement(dnsoption, "number").text = '6'
			ET.SubElement(dnsoption, "value").text = DNS1 + ',' + DNS2
			ET.SubElement(dnsoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Policy - OptionItems - NTPServers
			try:
				NTP1 = rowdata.get('NTP1').encode('utf-8')
				NTP2 = rowdata.get('NTP2').encode('utf-8')
			except:
				NTP1 = rowdata.get('NTP1')
				NTP1 = str(NTP1)
				NTP2 = rowdata.get('NTP2')
				NTP2 = str(NTP2)
			ntpoption = ET.SubElement (optionlist, "OptionItem")
			ET.SubElement(ntpoption, "number").text = '42'
			ET.SubElement(ntpoption, "value").text = NTP1 + ',' + NTP2
			ET.SubElement(ntpoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Policy - OptionItems - AddressTime
			try:
				LeaseDuration = rowdata.get('Lease Duration').encode('utf-8')
			except:
				LeaseDuration = rowdata.get('Lease Duration')
				LeaseDuration = str(LeaseDuration)
			ldoption = ET.SubElement (optionlist, "OptionItem")
			ET.SubElement(ldoption, "number").text = '51'
			ET.SubElement(ldoption, "value").text = LeaseDuration
			ET.SubElement(ldoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Policy - OptionItems - DomainSearchName
			try:
				DomainName = rowdata.get('Domain Name').encode('utf-8')
			except:
				DomainName = rowdata.get('Domain Name')
				DomainName = str(DomainName)
			dnsearchoption = ET.SubElement (optionlist, "OptionItem")
			ET.SubElement(dnsearchoption, "number").text = '15'
			ET.SubElement(dnsearchoption, "value").text = DomainName
			ET.SubElement(dnsearchoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Policy - TenantID
			try:
				TenantID = rowdata.get('TenantID').encode('utf-8')
			except:
				TenantID = rowdata.get('TenantID')
				TenantID = str(TenantID)
			ET.SubElement(policyroot, "TenantID").text = TenantID
			# SAVE FILE
			xmldata = ET.tostring(policyroot)
			xmldata = xml.dom.minidom.parseString(xmldata)
			xmldata = xmldata.toprettyxml()
			if 'e' in ImportTypeq.lower():
				filepath = ExportLocation + ZoneName + 'policy.xml'
				myfile = open(filepath, "w")  
				myfile.write(xmldata) 
				myfile.close()
			# Network Registrar API
			if 'i' in ImportTypeq.lower():
				try:
					r = requests.put(ZonePolicyURL, data=xmldata, headers=headers, auth=requests.auth.HTTPBasicAuth(NRUSERq,NRPASSWORDq), verify=False);
					status_code = r.status_code
					if (status_code == 429):
						print "API is currently being rate-limited. Pausing for 60 seconds."
						time.sleep(60)
						r = requests.put(ZonePolicyURL, data=xmldata, headers=headers, auth=requests.auth.HTTPBasicAuth(NRUSERq,NRPASSWORDq), verify=False);
						status_code = r.status_code
					resp = r.text
					if status_code == 201 or status_code == 202:
						print ("The following row of data was successfully imported: " + ZoneName)
					else :
						r.raise_for_status()
						print ("Error occurred in importing the following row: " + ZoneName + " Post error was " +resp)
				except requests.exceptions.HTTPError as err:
					if status_code == 400:
						print "Data " + ZoneName + " might already exist. Error code 400"
					else :
						print ("Error in connection to the server: "+str(err))	
				finally:
					if r : r.close()
		## ZONE POLICY END ####
		#                     #
		## DHCP SCOPE START ###
		if "y" in scopeq.lower():
			scoperoot = ET.Element("Scope")
			# Scope - FirstAvailable
			ET.SubElement(scoperoot, "allocateFirstAvailable").text = 'true'
			# Scope - BackupPercentage
			ET.SubElement(scoperoot, "backupPct").text = '15%'
			# Scope - BootP
			ET.SubElement(scoperoot, "bootp").text = 'disabled'
			# Scope - DHCP
			ET.SubElement(scoperoot, "dhcp").text = 'enabled'
			# Scope - ActivationStatus
			ET.SubElement(scoperoot, "deactivated").text = 'disabled'
			# Scope - Name
			ET.SubElement(scoperoot, "name").text = ZoneName
			# Scope - pingClients
			ET.SubElement(scoperoot, "pingClients").text = 'enabled'
			# Scope - pingTimeout
			ET.SubElement(scoperoot, "pingTimeout").text = '1000'
			# Scope - rangeList
			rangelist = ET.SubElement(scoperoot, "rangeList")
			# Scope - RangeItem
			rangeitem = ET.SubElement(rangelist, "RangeItem")
			try:
				networkaddress = rowdata.get('Network').encode('utf-8')
				networksubnet = rowdata.get('Subnet Mask').encode('utf-8')
			except:
				networkaddress = rowdata.get('Network')
				networkaddress = str(networkaddress)
				networksubnet = rowdata.get('Subnet Mask')
				networksubnet = str(networksubnet)
			networkcidr = IPAddress(networksubnet).netmask_bits()
			networkfullsubnet = str(networkaddress) + '/' + str(networkcidr)
			zonenetwork = IPNetwork(networkfullsubnet)
			zonenetworkdefault = zonenetwork.network
			zonenetworksize = zonenetwork.size
			zonenetworkstr = str(zonenetworkdefault)
			zonenetworkstr = zonenetworkstr.split('.')
			firstoctet = zonenetworkstr[0]
			secondoctet = zonenetworkstr[1]
			thirdoctet = zonenetworkstr[2]
			lastoctet = zonenetworkstr[3]
			dhcpfirstusable = int(lastoctet) + 10
			dhcpfirstusable = str(firstoctet) + '.' + str(secondoctet) + '.' + str(thirdoctet) + '.' + str(dhcpfirstusable)
			dhcplastusable = int(lastoctet) + zonenetworksize - 2
			dhcplastusable = str(firstoctet) + '.' + str(secondoctet) + '.' + str(thirdoctet) + '.' + str(dhcplastusable)
			ET.SubElement(rangeitem, "end").text = dhcplastusable
			ET.SubElement(rangeitem, "start").text = dhcpfirstusable
			# Scope - Subnet
			try:
				networkaddress = rowdata.get('Network').encode('utf-8')
				networksubnet = rowdata.get('Subnet Mask').encode('utf-8')
			except:
				networkaddress = rowdata.get('Network')
				networkaddress = str(networkaddress)
				networksubnet = rowdata.get('Subnet Mask')
				networksubnet = str(networksubnet)
			networkcidr = IPAddress(networksubnet).netmask_bits()
			networkfullsubnet = str(networkaddress) + '/' + str(networkcidr)
			ET.SubElement(scoperoot, "subnet").text = networkfullsubnet
			# Scope - tenantID
			try:
				TenantID = rowdata.get('TenantID').encode('utf-8')
			except:
				TenantID = rowdata.get('TenantID')
				TenantID = str(TenantID)
			ET.SubElement(scoperoot, "tenantId").text = TenantID
			# Scope - VpnID
			try:
				VpnID = rowdata.get('VpnID').encode('utf-8')
			except:
				VpnID = rowdata.get('VpnID')
				VpnID = str(TenantID)
			ET.SubElement(scoperoot, "vpnId").text = VpnID
			# Scope - Policy
			try:
				scopepolicyvar = rowdata.get('Scope Policy').encode('utf-8')
			except:
				scopepolicyvar = rowdata.get('Scope Policy')
				scopepolicyvar = str(scopepolicyvar)
			ET.SubElement(scoperoot, "policy").text = scopepolicyvar
			# Scope - embeddedPolicy
			embeddedPolicy = ET.SubElement(scoperoot, "embeddedPolicy")
			# Scope - embeddedPolicy - GracePeriod
			try:
				gracePeriod = rowdata.get('GracePeriod').encode('utf-8')
			except:
				gracePeriod = rowdata.get('GracePeriod')
				gracePeriod = str(gracePeriod)
			ET.SubElement(embeddedPolicy, "gracePeriod").text = gracePeriod
			# Scope - embeddedPolicy - Name
			try:
				scopetemplatevar = rowdata.get('Scope Template').encode('utf-8')
			except:
				scopetemplatevar = rowdata.get('Scope Template')
				scopetemplatevar = str(scopetemplatevar)
			scopetemplatevar = 'scope-template-policy:' + scopetemplatevar
			ET.SubElement(embeddedPolicy, "name").text = scopetemplatevar
			# Scope - embeddedPolicy - OfferTimeout
			try:
				offerTimeout = rowdata.get('offerTimeout').encode('utf-8')
			except:
				offerTimeout = rowdata.get('offerTimeout')
				offerTimeout = str(offerTimeout)
			ET.SubElement(embeddedPolicy, "offerTimeout").text = offerTimeout
			# Scope - embeddedPolicy - OptionList
			dhcpoptionlist = ET.SubElement(embeddedPolicy, "optionList")
			# Scope - embeddedPolicy - OptionItems - Gateway
			try:
				Gateway = rowdata.get('Gateway').encode('utf-8')
			except:
				Gateway = rowdata.get('Gateway')
				Gateway = str(Gateway)
			gatewayoption = ET.SubElement (dhcpoptionlist, "OptionItem")
			ET.SubElement(gatewayoption, "number").text = '3'
			ET.SubElement(gatewayoption, "value").text = Gateway
			ET.SubElement(gatewayoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - OptionItems - DNSServer
			try:
				DNS1 = rowdata.get('DNS1').encode('utf-8')
				DNS2 = rowdata.get('DNS2').encode('utf-8')
			except:
				DNS1 = rowdata.get('DNS1')
				DNS1 = str(DNS1)
				DNS2 = rowdata.get('DNS2')
				DNS2 = str(DNS2)
			dnsoption = ET.SubElement (dhcpoptionlist, "OptionItem")
			ET.SubElement(dnsoption, "number").text = '6'
			ET.SubElement(dnsoption, "value").text = DNS1 + ',' + DNS2
			ET.SubElement(dnsoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - OptionItems - NTPServers
			try:
				NTP1 = rowdata.get('NTP1').encode('utf-8')
				NTP2 = rowdata.get('NTP2').encode('utf-8')
			except:
				NTP1 = rowdata.get('NTP1')
				NTP1 = str(NTP1)
				NTP2 = rowdata.get('NTP2')
				NTP2 = str(NTP2)
			ntpoption = ET.SubElement (dhcpoptionlist, "OptionItem")
			ET.SubElement(ntpoption, "number").text = '42'
			ET.SubElement(ntpoption, "value").text = NTP1 + ',' + NTP2
			ET.SubElement(ntpoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - OptionItems - AddressTime
			try:
				LeaseDuration = rowdata.get('Lease Duration').encode('utf-8')
			except:
				LeaseDuration = rowdata.get('Lease Duration')
				LeaseDuration = str(LeaseDuration)
			ldoption = ET.SubElement (dhcpoptionlist, "OptionItem")
			ET.SubElement(ldoption, "number").text = '51'
			ET.SubElement(ldoption, "value").text = LeaseDuration
			ET.SubElement(ldoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - embeddedPolicy - OptionItems - DomainSearchName
			try:
				DomainName = rowdata.get('Domain Name').encode('utf-8')
			except:
				DomainName = rowdata.get('Domain Name')
				DomainName = str(DomainName)
			dnsearchoption = ET.SubElement (dhcpoptionlist, "OptionItem")
			ET.SubElement(dnsearchoption, "number").text = '15'
			ET.SubElement(dnsearchoption, "value").text = DomainName
			ET.SubElement(dnsearchoption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - embeddedPolicy - OptionItems - Option150
			try:
				Option150 = rowdata.get('Option150').encode('utf-8')
			except:
				Option150 = rowdata.get('Option150')
				Option150 = str(Option150)
			if not 'None' in Option150:
				tftpserveroption = ET.SubElement (dhcpoptionlist, "OptionItem")
				ET.SubElement(tftpserveroption, "number").text = '150'
				ET.SubElement(tftpserveroption, "value").text = Option150
				ET.SubElement(tftpserveroption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - embeddedPolicy - OptionItems - Option43
			try:
				Option43 = rowdata.get('Option43').encode('utf-8')
			except:
				Option43 = rowdata.get('Option43')
				Option43 = str(Option43)
			if not 'None' in Option43:
				tftpserveroption = ET.SubElement (dhcpoptionlist, "OptionItem")
				ET.SubElement(tftpserveroption, "number").text = '43'
				ET.SubElement(tftpserveroption, "value").text = Option43
				ET.SubElement(tftpserveroption, "optionDefinitionSetName").text = 'dhcp-config'
			# Scope - embeddedPolicy - embeddedPolicy - OptionItems - Option60
			try:
				Option60 = rowdata.get('Option60').encode('utf-8')
			except:
				Option60 = rowdata.get('Option60')
				Option60 = str(Option60)
			if not 'None' in Option60:
				tftpserveroption = ET.SubElement (dhcpoptionlist, "OptionItem")
				ET.SubElement(tftpserveroption, "number").text = '60'
				ET.SubElement(tftpserveroption, "value").text = Option60
				ET.SubElement(tftpserveroption, "optionDefinitionSetName").text = 'dhcp-config'
			# Make XML pretty
			xmldata = ET.tostring(scoperoot)
			xmldata = xml.dom.minidom.parseString(xmldata)
			xmldata = xmldata.toprettyxml()
			# SAVE FILE
			if 'e' in ImportTypeq.lower():
				filepath = ExportLocation + ZoneName + 'scope.xml'
				myfile = open(filepath, "w")  
				myfile.write(xmldata) 
				myfile.close()
			# Network Registrar API
			if 'i' in ImportTypeq.lower():
				try:
					r = requests.put(ZoneScopeURL, data=xmldata, headers=headers, auth=requests.auth.HTTPBasicAuth(NRUSERq,NRPASSWORDq), verify=False);
					status_code = r.status_code
					if (status_code == 429):
						print "API is currently being rate-limited. Pausing for 60 seconds."
						time.sleep(60)
						r = requests.put(ZoneScopeURL, data=xmldata, headers=headers, auth=requests.auth.HTTPBasicAuth(NRUSERq,NRPASSWORDq), verify=False);
						status_code = r.status_code
					resp = r.text
					if status_code == 200 or status_code == 202 or status_code == 201:
						print ("The following row of data was successfully imported: " + ZoneName)
					else :
						r.raise_for_status()
						print ("Error occurred in importing the following row: " + ZoneName + " Post error was " +resp)
				except requests.exceptions.HTTPError as err:
					if status_code == 400:
						print "Data " + ZoneName + " might already exist. Error code 400"
					else :
						print ("Error in connection to the server: "+str(err))	
				finally:
					if r : r.close()
	except:
		print 'Skipping row ' + ZoneName
print 'Completed Script'
# Exit